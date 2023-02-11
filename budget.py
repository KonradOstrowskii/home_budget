import pandas as pd
import os
import openpyxl


class Budget:
    def __init__(self, categories: list):
        """
        __init__  Creating list with categories

        Args:
            categories (list): List of categories
        """
        self.categories = categories
        self.data = {}
        for category in self.categories:
            self.data[category] = 0
            

    def add(self, category: str, amount: float):
        """
        Feature to add expenses for `category`
        Save file as `xlsx` file

        Args:
            category (list): `category` in the `categories` list
            amount (float): Amount entered by User
        """
        self.data[category] = amount
        Budget.save(self,'budget.xlsx')
        
    def remove(self, category : str, amount: float):
        """
        Feature to remove expenses from `category`
        Save file as `xlsx` file

        Args:
            category (list): `category` in the `categories` list
            amount (float): Amount entered by User
        """
        self.data[category] = -amount
        Budget.save(self,'budget.xlsx')
    
    def save(self, file_name : object) -> object:
        """
        1.Load the existing data from the Excel file
        2.Update the values in the existing row
        3.Save the changes to the Excel file

        Args:
            file_name (object): `xlsx` file
        """
        # Load the existing data from the Excel file
        wb = openpyxl.load_workbook(file_name)
        sheet = wb.active
        last_row = sheet.max_row

        # Update the values in the existing row
        for i, category in enumerate(self.categories):
            sheet.cell(row=last_row, column=i+1).value = self.data[category]

        # Save the changes to the Excel file
        sheet.insert_rows(sheet.max_row + 1)
        wb.save(file_name)
        
    def new_row(self,file_name: object ="budget.xlsx") -> object:
        """
        1.Load the existing data from the Excel file
        2.Find the last row of the sheet
        3.Update the values in the existing row
        
        Args:
            file_name (object, optional): By default to `"budget.xlsx"`.

        Returns:
            object: Return updated xlsx file
        """
        # Load the existing data from the Excel file
        wb = openpyxl.load_workbook(file_name)
        sheet = wb.active

        # Find the last row of the sheet
        last_row = sheet.max_row

        # Update the values in the existing row
        for i, category in enumerate(self.categories):
            sheet.cell(row=last_row+1, column=i+1).value = self.data[category]
        wb.save(file_name)
        
    def open(self,file_name: str) -> object:
        """
        1.Find path to file if exists open it, else create file in indicated path
        2.Create a DataFrame with the budget data
        3.Drop rows with 0 values
        4.Save the DataFrame to the `Excel` file
        5.Save the combined data to the `Excel` file

        Returns:
            object: excel file`xlsx`
        """
        folder_name = "users_budgets"

        folder_path = os.path.join(os.getcwd(), folder_name)

        try:
            os.makedirs(folder_path)
        except OSError:
            print ("Creation of the directory %s failed because folder already exist " % folder_path)
        else:
            print ("Successfully created the directory %s" % folder_path)       
            # try:
            print("FileNotFoundError")
            print("Creating New File")
            # Create a DataFrame with the budget data
            df = pd.DataFrame([self.data])
            
            # Dropping rows with 0 values
            df.drop(df.index[0], inplace=True)
                
            # Save the DataFrame to the Excel file
            df.to_excel(file_name, index=False)
            
            # Save the combined data to the Excel file
            df.style.set_caption('Budget') \
                .format({col: '${:,.2f}'.format for col in self.categories}) \
                .hide(axis='index') \
                .to_excel(file_name, index=False)
            # except TypeError as e:
            #     print("'builtin_function_or_method' object is not iterable")
                
    def budget_category(self,budget_category_show: list)-> str:
        """
        Function to enumerate all categories in the list

        Args:
            budget_category_show (list): Enumerated list of categories

        Returns:
            str: Return enumerated elements of the list
        """
        for number , category in enumerate(budget_category_show):
                return(f"{number+1}:{category}")
                
    def transaction_history(self: object)-> object:
        """
        Function showing transaction history

        Args:
            self (object): file `xlsx`

        Returns:
            object: Return transaction of history
        """
        fd = pd.read_excel('budget.xlsx')
        print(fd)
        
    def count(self,budget_to_count: list) -> list:
        """
        Function counting amount from all category

        Args:
            budget_to_count (list): list of category to count

        Returns:
            list: Returning list of category and sum of all amounts
        """
        fd = pd.read_excel('budget.xlsx')
        # Budget.budget_category(self)
        try:
            for number , category in enumerate(budget_to_count):
                print(f"{number+1}:{category}")
                sum =fd[category].sum()
                print(sum)
        except KeyError as e:
                # This code will be executed if a KeyError occurs
                print("Error: key not found in budget list")
                
